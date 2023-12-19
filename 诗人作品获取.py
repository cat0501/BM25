import requests
from bs4 import BeautifulSoup
import openpyxl
from selenium import webdriver
import re
import json
from datetime import datetime

requests.adapters.DEFAULT_RETRIES = 5  # 增加重连次数

login_data = {
    '_ZVING_METHOD': 'member.login',
    '_ZVING_URL': '%2Fzcms%2Fmember%2Flogin',
    '_ZVING_DATA': '{"SiteID":"122","Referer":"https://www.zgshige.com/zcms/poem/list?SiteID=122&poetname=%E7%8E%99%E7%92%A0&articleID=1364604&articleContributeUID=33912&catalogID=15111","UserName":"Gbo99P9cwhD01/bGY7/n9A==","Password":"gkAd52S7uZP6Ciy+gUtzbA==","AuthCode1":"","LoginStatusStand":"Y","AuthCode":"","key":"aecd5761d6f94f74"}',
    '_ZVING_DATA_FORMAT': 'json'
}

def getHTMLText(url):
    try:
        kv = {'user-agent': 'Mozilla/5.0'}
        r = requests.get(url, timeout=30)

        r.raise_for_status()  # 如果返回值不是200触发HTTP异常
        r.encoding = r.apparent_encoding  # 改变爬取的网页编码，使得print结果不是乱码
        return r.text  # 获取网页的所有信息
    except:
        return "Error!"

def getHTMLText2(url):
    try:
        session = requests.Session()
        response = session.post('https://www.zgshige.com/zcms/ajax/invoke', data=login_data)
        if response.status_code == 200:  # 判断登录成功
            # 重新加载首页
            response = session.get(url)
        else:
            print("登录失败")
        response.encoding = response.apparent_encoding
        return response.text  # 获取网页的所有信息
    except:
        return "Error!"


def extract_info(page_content, url):
    try:
        soup = BeautifulSoup(page_content, 'html.parser')

        # 诗人姓名
        name_element = soup.find('div', class_='sr_name')
        author_name = name_element.text.strip() if name_element else ""
        # 作品名称
        avatar_elements = soup.find_all('p', class_='sr_dt_title')
        poem_names = []
        urls = []
        for avatar_element in avatar_elements:
            title_element = avatar_element.find('a')
            poem_name = title_element.text.strip() if title_element else ""
            poem_names.append(poem_name)
            print("Poem Name:", poem_name)

            url = title_element.get('href') if title_element else ""
            urls.append(url)

        driver.get(url)

        # 简介
        introduction_element = soup.select_one('.b-a .p-sm p')
        introduction = introduction_element.text.strip() if introduction_element else ""


    except Exception as e:
        print("Error: {}".format(e))
        return {"Error": str(e)}

    return {'author_name': author_name, 'poem_names': poem_names, 'urls': urls,
            'introduction': introduction}


# page_content 诗歌页面内容
def get_poem_html_info(page_content, url):
    try:
        soup = BeautifulSoup(page_content, 'html.parser')
        # 时间
        time_element = soup.find('span', class_='p-l-sm p-r-sm')
        time = time_element.text.strip() if time_element else ''

        # 内容
        content_element = soup.find('div', class_='m-lg font14 mwebfont')
        content = content_element.text.strip() if content_element else ''

        # 浏览量
        pattern1 = re.compile("^hitcount")
        hit_element = soup.find('span', id=pattern1)
        hit_value = hit_element.text.strip() if hit_element else ''

        # 点赞量
        favor_value_content = 0
        good_and_bad_div = soup.find("div", {"class": "goodAndBad"})
        script_tags = good_and_bad_div.find_all("script", text=re.compile(r'\$.getScript\(".*"\)'))
        if script_tags:
            first_script_tag = script_tags[0]
            value = re.search(r'\$\.(getScript)\("(.*)"\);', first_script_tag.string)
            if value:
                get_script_value = value.group(2)
                print("--------------点赞数量url开始-----------------")
                print(get_script_value)
                favor_value_content = getHTMLText(get_script_value)
                # 使用正则表达式来找到 innerHTML 的值
                match = re.search(r'innerHTML=(\d+)', favor_value_content)
                if match:
                    inner_html_value = match.group(1)
                    favor_value_content = inner_html_value
                    print(inner_html_value)

        # 评论量
        # 在整个源代码中搜索函数loadActivityConfig
        var_dc = ''
        script_tags = soup.find_all("script")
        for script in script_tags:
            script_content = script.string
            if script_content and "function loadActivityConfig()" in script_content:
                print("loadActivityConfig function found!")
                var_dc = script_content
                break
        # 使用正则表达式提取var dc的内容
        var_dc_match = re.search(r"var dc = {(.*?)};", var_dc, re.DOTALL)
        if var_dc_match:
            var_dc_content = var_dc_match.group(1)
            print("var dc: ", var_dc_content)

        # 使用正则表达式提取ContentID和CatalogID的值
        content_id_match = re.search(r'ContentID:"(.*?)"', var_dc_content)
        catalog_id_match = re.search(r'CatalogID:"(.*?)"', var_dc_content)

        if content_id_match:
            content_id = content_id_match.group(1)
            print("ContentID: ", content_id)

        if catalog_id_match:
            catalog_id = catalog_id_match.group(1)
            print("CatalogID: ", catalog_id)

        comment_url = "https://www.zgshige.com/zcms/poetrylist/getcommonlist?ContentID=" + content_id + "&ContentType=Poetry&CatalogID=" + catalog_id + "&"
        comment_content = getHTMLText2(comment_url)
        # 解析返回的 JSON 数据
        data = json.loads(comment_content)
        # 统计评论数量
        comment_count = len(data['data'])
        print("评论数量: ", comment_count)

        driver.get(url)







    except Exception as e:
        print("Error: {}".format(e))
        return {"Error": str(e)}

    return {'time': time, 'content': content, 'hit_value': hit_value, 'favor_value_content': favor_value_content,
            'comment_count': comment_count}


def crawl_info(url_list, driver):
    # 打开excel文件
    workbook = openpyxl.load_workbook('D://work//test//in.xlsx')
    sheet = workbook.active

    sheet['A1'] = '姓名'
    sheet['B1'] = '诗歌名称'
    sheet['C1'] = '诗歌地址'
    sheet['D1'] = '时间'
    sheet['E1'] = '内容'
    sheet['F1'] = '浏览量'
    sheet['G1'] = '点赞量'
    sheet['H1'] = '评论量'
    # TODO 完善

    # DEBUG writed 写入行号有问题
    writed = 2
    # url_list是诗人作品列表地址集合
    for i, url in enumerate(url_list, start=2):
        print("-------------------------------------------------------: ", i)
        print("当前时间为:", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        num_poems = 0
        try:
            # 请求 url
            page_content = getHTMLText(url)
            # 返回数据处理
            info = extract_info(page_content, url)

            poem_names = info['poem_names']
            num_poems = len(poem_names)  # 获取 poem_names 列表长度

            urls = info['urls']
            num_urls = len(urls)

            for j, poem_name in enumerate(poem_names, start=0):
                poem_info_content = getHTMLText2(urls[j])
                poem_html_info = get_poem_html_info(poem_info_content, url)

                sheet.cell(row=writed + j, column=1, value=info['author_name'])
                sheet.cell(row=writed + j, column=2, value=poem_name)
                sheet.cell(row=writed + j, column=3, value=urls[j])
                sheet.cell(row=writed + j, column=4, value=poem_html_info['time'])
                sheet.cell(row=writed + j, column=5, value=poem_html_info['content'])
                sheet.cell(row=writed + j, column=6, value=poem_html_info['hit_value'])
                sheet.cell(row=writed + j, column=7, value=poem_html_info['favor_value_content'])
                sheet.cell(row=writed + j, column=8, value=poem_html_info['comment_count'])
        except Exception as e:
            print(f'在爬取URL：{url}时出现错误：{e}')

        writed += num_poems  # 递增行索引，跳过已写入的行

    workbook.save('poem_5.xlsx')


# 从Excel中读取URL列表
def read_url_list_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    url_list = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        url = row[1]
        if url:
            url_list.append(url)

    return url_list


if __name__ == '__main__':
    url_list = read_url_list_from_excel('D://learn//毕业设计//BM25-master//poet_url1211_02.xlsx')  # 读取URL列表的文件路径
    driver = webdriver.Chrome()

    try:
        crawl_info(url_list, driver)  # 将driver传递给crawl_info函数
    finally:
        driver.quit()  # 确保在函数执行结束或发生异常时关闭driver
