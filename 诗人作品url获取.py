import requests
from bs4 import BeautifulSoup
import openpyxl
from selenium import webdriver
# 启动 Chrome 浏览器
from selenium.webdriver.common.by import By
import re
from urllib.parse import quote, unquote

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

requests.adapters.DEFAULT_RETRIES = 5 # 增加重连次数

def getHTMLText(url):
    try:
        kv = {'user-agent': 'Mozilla/5.0'}
        r = requests.get(url, timeout=30)

        r.raise_for_status()  # 如果返回值不是200触发HTTP异常
        r.encoding = r.apparent_encoding  # 改变爬取的网页编码，使得print结果不是乱码
        return r.text  # 获取网页的所有信息
    except:
        return "Error!"

def extract_info(page_content,url):
    try:
        soup = BeautifulSoup(page_content, 'html.parser')

        #姓名
        name_element = soup.find('div', class_='sr_name')
        name = name_element.text.strip() if name_element else ""
        #头像
        avatar_element = soup.find('div', class_='sr_face_img')
        avatar = avatar_element.img['src'] if avatar_element and avatar_element.img else ""

        driver.get(url)
        #粉丝数
        fans_element = driver.find_element(By.ID,'poemPowder')
        fans = fans_element.text if fans_element else "0"

        #赞赏数
        likes_element = driver.find_element(By.ID, 'contentCount')

        likes =  re.findall(r'\d+', likes_element.text)# if likes_element else ""
        likes = int(likes[0]) if likes else "0"

        #作品数
        works_element = driver.find_element(By.ID, 'poetry')
        works = works_element.text if works_element else "0"

        #评论数
        comments_element = driver.find_element(By.CLASS_NAME, 'comment-persontotal')
        comments = comments_element.text if comments_element else "0"


        #简介
        introduction_element =  soup.select_one('.b-a .p-sm p')
        introduction = introduction_element.text.strip() if introduction_element else ""

        #作品地址url
        # 找到含有重定向链接的元素
        links = soup.find_all("a", class_="m-l-lg m-l-lg")

        # 确认链接列表只有一个元素
        encoded_href = ''
        if len(links) == 1:
            link = links[0]
            href = link.get("href")
            # 提取 https 后面的部分
            index = href.find("https")
            href = href[index:]
            href = href[: -2]

            print(href)
            encoded_href = quote(href, safe = ":/?&=")
            print(encoded_href)
        else:
            print("链接列表不唯一")




    except Exception as e:
        print("Error: {}".format(e))
        return {"Error": str(e)}

    return {'name': name, 'avatar': avatar, 'fans': fans, 'likes': likes, 'works': works, 'comments': comments,
            'introduction': introduction, 'encoded_href': encoded_href}


def crawl_info(url_list, driver):
    # 打开excel文件
    workbook = openpyxl.load_workbook('D://work//test//in.xlsx')
    sheet = workbook.active
    sheet['A1'] = '姓名'
    sheet['B1'] = '头像'
    sheet['C1'] = '粉丝数'
    sheet['D1'] = '赞赏数'
    sheet['E1'] = '作品数'
    sheet['F1'] = '评论数'
    sheet['G1'] = '简介'
    sheet['H1'] = '作品页面url'

    for i, url in enumerate(url_list, start=2):
        try:
            page_content = getHTMLText(url)
            info = extract_info(page_content,url)
            sheet.cell(row=i, column=1, value=info['name'])
            sheet.cell(row=i, column=2, value=info['avatar'])
            sheet.cell(row=i, column=3, value=info['fans'])
            sheet.cell(row=i, column=4, value=info['likes'])
            sheet.cell(row=i, column=5, value=info['works'])
            sheet.cell(row=i, column=6, value=info['comments'])
            sheet.cell(row=i, column=7, value=info['introduction'])
            sheet.cell(row=i, column=8, value=info['encoded_href'])

        except Exception as e:
            print(f'在爬取URL：{url}时出现错误：{e}')

    workbook.save('poet_output2.xlsx')

# 从Excel中读取URL列表
def read_url_list_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    url_list = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        url = row[1]
        if url:
            url_list.append(url)

    return url_list

if __name__=='__main__':
    url_list = read_url_list_from_excel('D://work//test//in03.xlsx')  # 读取URL列表的文件路径
    driver = webdriver.Chrome()

    try:
        crawl_info(url_list, driver)  # 将driver传递给crawl_info函数
    finally:
        driver.quit()  # 确保在函数执行结束或发生异常时关闭driver